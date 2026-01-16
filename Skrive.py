from openpyxl import load_workbook
import pandas as pd

def printxl(transactions, sheetname, start_cell, filename="russens-regnskap_helautomatisert15.xlsx",headers=True):


    try:
        wb = load_workbook(filename)
    except FileNotFoundError:
        from openpyxl import Workbook
        wb = Workbook()

    # Select or create sheet
    ws = wb[sheetname] if sheetname in wb.sheetnames else wb.create_sheet(sheetname)

    # Convert start_cell to row/column
    from openpyxl.utils.cell import coordinate_to_tuple
    start_row, start_col = coordinate_to_tuple(start_cell)
    
    # Write headers
    if headers == True:
        for j, col in enumerate(transactions.columns, start=start_col):
            ws.cell(row=start_row, column=j, value=col)

    # Write data
    for i, row in enumerate(transactions.itertuples(index=False), start=start_row+1):
        for j, val in enumerate(row, start=start_col):
            ws.cell(row=i, column=j, value=val)

    wb.save(filename)
    print(f"✅ Data written to '{filename}' → sheet '{sheetname}' starting at {start_cell}")
